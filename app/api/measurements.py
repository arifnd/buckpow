import csv
import io
from datetime import datetime, timezone, timedelta

from fastapi import APIRouter, Depends, HTTPException, Query, Request
from fastapi.responses import Response, StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from pydantic import BaseModel
from sqlalchemy.orm import Session

from app.database import get_db
from app.models import User
from app.models.session import Session as SessionModel
from app.services.measurement_service import MeasurementService
from app.services.device_service import DeviceService
from app.services.audit_service import AuditService
from app.utils.client_ip import get_client_ip
from app.dependencies import get_api_key_device, require_user
from app.api.health import MIN_FIRMWARE_VERSION
from app.schemas import MeasurementCreate

router = APIRouter()


def _parse_version(v):
    try:
        return tuple(int(x) for x in v.split('.'))
    except (ValueError, AttributeError):
        return None


@router.post('/measurements', status_code=201)
def receive_measurement(
    body: MeasurementCreate,
    db: Session = Depends(get_db),
    device=Depends(get_api_key_device),
):
    if device is None:
        device = DeviceService.get_by_device_id(db, body.device_id)
        if not device:
            device = DeviceService.get_or_create(db, body.device_id)
        if not device.enabled:
            raise HTTPException(status_code=403, detail='Device is disabled')
    else:
        if body.device_id != device.device_id:
            raise HTTPException(status_code=403, detail='device_id does not match the authenticated device')

    fw = body.firmware_version or ''
    outdated = False
    if fw:
        parsed = _parse_version(fw)
        min_fw = _parse_version(MIN_FIRMWARE_VERSION)
        if parsed and min_fw and parsed < min_fw:
            outdated = True
        if fw != device.firmware_version:
            device.firmware_version = fw
            db.commit()
    elif not device.firmware_version:
        device.firmware_version = 'unknown'
        db.commit()

    try:
        measurement = MeasurementService.create(
            db,
            device_id_str=body.device_id,
            bus_voltage=float(body.bus_voltage),
            shunt_voltage=float(body.shunt_voltage),
            current=float(body.current),
            power=float(body.power),
        )
        resp = {'status': 'success', 'id': measurement.id}
        if outdated:
            from fastapi.responses import JSONResponse
            resp_obj = JSONResponse(resp, status_code=201)
            resp_obj.headers['X-Firmware-Outdated'] = 'true'
            return resp_obj
        return resp
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get('/measurements')
def get_measurements(
    device_id: int | None = Query(None),
    session_id: int | None = Query(None),
    page: int = Query(1),
    per_page: int = Query(50),
    start_date: str | None = Query(None),
    end_date: str | None = Query(None),
    db: Session = Depends(get_db),
    _current_user: User = Depends(require_user),
):
    pagination = MeasurementService.get_paginated(
        db, page=page, per_page=per_page,
        device_id=device_id, session_id=session_id,
        start_date=start_date, end_date=end_date,
    )
    return {
        'measurements': [m.to_dict() for m in pagination.items],
        'page': pagination.page,
        'pages': pagination.pages,
        'total': pagination.total,
        'per_page': pagination.per_page,
    }


@router.get('/measurements/export/csv')
def export_csv(
    request: Request,
    device_id: int | None = Query(None),
    session_id: int | None = Query(None),
    start_date: str | None = Query(None),
    end_date: str | None = Query(None),
    db: Session = Depends(get_db),
    _current_user: User = Depends(require_user),
):
    rows = MeasurementService.get_all_filtered(
        db, device_id=device_id, session_id=session_id,
        start_date=start_date, end_date=end_date,
    )
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['ID', 'Device', 'Session', 'Bus Voltage', 'Shunt Voltage', 'Load Voltage', 'Current (A)', 'Power (W)', 'Energy (Wh)', 'Timestamp'])
    for m in rows:
        writer.writerow([
            m.id,
            m.device_ref.device_id if m.device_ref else '',
            m.session_ref.name if m.session_ref else '',
            m.bus_voltage, m.shunt_voltage, m.load_voltage,
            m.current, m.power, m.energy,
            m.created_at.isoformat() if m.created_at else '',
        ])
    ip = get_client_ip(request)
    AuditService.log(db, 'export.csv', user_id=_current_user.id, target_type='export', ip_address=ip, details={'rows': len(rows)})
    filename = 'measurements.csv'
    if session_id:
        session = db.get(SessionModel, session_id)
        if session:
            safe_name = ''.join(c if c.isalnum() or c in ' -_' else '' for c in session.name).strip().replace(' ', '_')
            filename = f'{safe_name}_report.csv'
    return Response(
        content=output.getvalue(),
        media_type='text/csv',
        headers={'Content-Disposition': f'attachment; filename={filename}'},
    )


@router.get('/measurements/export/xlsx')
def export_xlsx(
    request: Request,
    device_id: int | None = Query(None),
    session_id: int | None = Query(None),
    start_date: str | None = Query(None),
    end_date: str | None = Query(None),
    db: Session = Depends(get_db),
    _current_user: User = Depends(require_user),
):
    rows = MeasurementService.get_all_filtered(
        db, device_id=device_id, session_id=session_id,
        start_date=start_date, end_date=end_date,
    )
    wb = Workbook()
    ws = wb.active
    ws.title = 'Measurements'

    headers = ['ID', 'Device', 'Session', 'Bus Voltage', 'Shunt Voltage', 'Load Voltage', 'Current (A)', 'Power (W)', 'Energy (Wh)', 'Timestamp']
    bold = Font(bold=True)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = bold

    for i, m in enumerate(rows, 2):
        ws.cell(row=i, column=1, value=m.id)
        ws.cell(row=i, column=2, value=m.device_ref.device_id if m.device_ref else '')
        ws.cell(row=i, column=3, value=m.session_ref.name if m.session_ref else '')
        ws.cell(row=i, column=4, value=m.bus_voltage)
        ws.cell(row=i, column=5, value=m.shunt_voltage)
        ws.cell(row=i, column=6, value=m.load_voltage)
        ws.cell(row=i, column=7, value=m.current)
        ws.cell(row=i, column=8, value=m.power)
        ws.cell(row=i, column=9, value=m.energy)
        ws.cell(row=i, column=10, value=m.created_at.isoformat() if m.created_at else '')

    for col in range(1, len(headers) + 1):
        max_len = 0
        for row in ws.iter_rows(min_col=col, max_col=col, values_only=True):
            val = str(row[0] or '')
            max_len = max(max_len, len(val))
        ws.column_dimensions[get_column_letter(col)].width = max_len + 3

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    ip = get_client_ip(request)
    AuditService.log(db, 'export.xlsx', user_id=_current_user.id, target_type='export', ip_address=ip, details={'rows': len(rows)})
    return Response(
        content=output.read(),
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment; filename=measurements.xlsx'},
    )


@router.get('/chart')
def chart_data(
    device_id: int | None = Query(None),
    session_id: int | None = Query(None),
    limit: int = Query(500),
    granularity: str | None = Query(None),
    range: str | None = Query(None, alias='range'),
    start_date: str | None = Query(None),
    end_date: str | None = Query(None),
    db: Session = Depends(get_db),
    _current_user: User = Depends(require_user),
):
    if granularity not in (None, 's', 'm', 'h', 'd'):
        granularity = None

    if range == '1h':
        start_date = datetime.now(timezone.utc) - timedelta(hours=1)
    elif range == '24h':
        start_date = datetime.now(timezone.utc) - timedelta(hours=24)
    elif range == '7d':
        start_date = datetime.now(timezone.utc) - timedelta(days=7)
    elif range == '30d':
        start_date = datetime.now(timezone.utc) - timedelta(days=30)

    data = MeasurementService.get_chart_data(
        db, limit=limit, device_id=device_id, session_id=session_id,
        granularity=granularity, start_date=start_date, end_date=end_date,
    )
    return data
